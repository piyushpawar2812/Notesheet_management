from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.db.models import Sum, Q
from openpyxl import Workbook, load_workbook
from .models import InventoryItem


# =========================================
# INVENTORY LIST
# =========================================

def inventory_stock(request):
    """Display inventory list with search and pagination"""
    
    search = request.GET.get('search', '')
    
    # Get all inventory items with notesheet relation
    data = InventoryItem.objects.select_related('notesheet').all().order_by('-id')
    
    # Apply search filter if provided
    if search:
        data = data.filter(
            Q(item_name__icontains=search) |
            Q(supplier_name__icontains=search) |
            Q(notesheet__notesheet_no__icontains=search) |
            Q(description__icontains=search)  # 🔥 Description पर भी search
        )
    
    # Pagination
    paginator = Paginator(data, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics for dashboard
    total_items = InventoryItem.objects.count()
    available_items = InventoryItem.objects.filter(is_available=True).count()
    out_of_stock_items = InventoryItem.objects.filter(is_available=False).count()
    total_value = InventoryItem.objects.aggregate(total=Sum('price'))['total'] or 0
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'total_items': total_items,
        'available_items': available_items,
        'out_of_stock_items': out_of_stock_items,
        'total_value': total_value,
    }
    
    return render(request, 'services/inventory_stock.html', context)


# =========================================
# EXPORT EXCEL
# =========================================

def export_inventory_excel(request):
    """Export all inventory data to Excel"""
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Inventory'
    
    # Headers with Description
    headers = [
        'Notesheet No',
        'Item Name',
        'Quantity',
        'Price (₹)',
        'Supplier Name',
        'Received Date',
        'Status',
        'Description'  # ✅ Description column added
    ]
    sheet.append(headers)
    
    # Data
    data = InventoryItem.objects.select_related('notesheet').all()
    
    for item in data:
        sheet.append([
            item.notesheet.notesheet_no if item.notesheet else 'N/A',
            item.item_name,
            item.quantity,
            float(item.price),
            item.supplier_name or 'N/A',
            str(item.added_date) if item.added_date else 'N/A',
            'Available' if item.is_available else 'Out of Stock',
            item.description or 'N/A'  # ✅ Description added
        ])
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=inventory_stock.xlsx'
    workbook.save(response)
    
    return response


# =========================================
# IMPORT EXCEL
# =========================================

def import_inventory_excel(request):
    """Import inventory data from Excel file with Description"""
    
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('inventory_stock')
    
    excel_file = request.FILES.get('excel_file')
    
    if not excel_file:
        messages.error(request, 'Please upload an Excel file.')
        return redirect('inventory_stock')
    
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
        return redirect('inventory_stock')
    
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        
        imported_count = 0
        skipped_count = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            # ✅ Description included (index 5)
            InventoryItem.objects.create(
                item_name=row[0] or 'Unknown',
                quantity=row[1] or 0,
                price=row[2] or 0.00,
                supplier_name=row[3] or '',
                added_date=row[4] if row[4] else None,
                description=row[5] or '',  # ✅ Description added
                is_available=True
            )
            imported_count += 1
        
        messages.success(
            request, 
            f'✅ Successfully imported {imported_count} items from Excel.'
        )
        
    except Exception as e:
        messages.error(request, f'❌ Error importing Excel: {str(e)}')
    
    return redirect('inventory_stock')